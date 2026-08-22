"""HSE daily-statistics workbook import and column definitions.

Mirrors the "Statistics" sheet of the project's daily HSE tracking workbook:
one row per calendar day, columns grouped into the same sections the source
spreadsheet uses (Administrative Data, Lagging Indicators, Leading
Indicators, Inspections and Audits, OHS Trainings, Permits).

This module has no Flask or database dependency — app.py owns storage and
routing, this module only knows how to read and validate the workbook.
"""

from __future__ import annotations

import io
import re
import zipfile
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


class ValidationError(Exception):
    """Raised when an uploaded workbook can't be safely imported."""


# (db_column, xlsx_header, aggregation, section, kind)
#   aggregation: "sum" | "avg" | "last"  (matches the workbook's own "PBI
#     Aggregation" row, so dashboard totals agree with the source sheet)
#   kind: "int" | "float" | "percent" (percent values are stored as a
#     fraction, e.g. 0.75, matching how the workbook itself stores them)
COLUMNS: list[tuple[str, str, str, str, str]] = [
    ("bec_manpower", "Total BEC Manpower", "avg", "Administrative Data", "float"),
    ("subcon_manpower", "Total Sub. Con Manpower", "avg", "Administrative Data", "float"),
    ("total_manpower", "Total Man Power", "avg", "Administrative Data", "float"),
    ("working_hours_per_day", "Working Hours / Day", "avg", "Administrative Data", "float"),
    ("total_man_hours", "Total Man Hours", "sum", "Administrative Data", "float"),
    ("safe_man_hours", "Hours Worked since last LTI (safe Man Hours)", "last", "Administrative Data", "float"),
    ("safety_population_required", "Safety Population Required", "avg", "Administrative Data", "int"),
    ("safety_population_available", "Safety Population Available", "avg", "Administrative Data", "int"),
    ("fatalities", "Number of Fatalities", "sum", "Lagging Indicators", "int"),
    ("lti_count", "Number of Lost Time Injuries (LTI)", "sum", "Lagging Indicators", "int"),
    ("ltifr", "Lost Time Injury Rate (LTIFR)", "last", "Lagging Indicators", "float"),
    ("fac_count", "First Aid Cases (FAC)", "sum", "Lagging Indicators", "int"),
    ("mtc_count", "Medical Treatment Cases (MTC)", "sum", "Lagging Indicators", "int"),
    ("pd_count", "Property Damages (PD)", "sum", "Lagging Indicators", "int"),
    ("fire_i_ii_count", "Fire Incident Level I & ii", "sum", "Lagging Indicators", "int"),
    ("fire_iii_count", "Fire Incidents Level iii", "sum", "Lagging Indicators", "int"),
    ("sic_count", "Security Incidents (SIC)", "sum", "Lagging Indicators", "int"),
    ("mvi_count", "Motor Vehicle Incidents (MVI)", "sum", "Lagging Indicators", "int"),
    ("trir", "Total Recordable Incidents Rate (TRIR)", "last", "Lagging Indicators", "float"),
    ("tri_count", "Total Recordable Incidents (LTI+Fire+MTC+PD+MVI)", "sum", "Lagging Indicators", "int"),
    ("near_miss_count", "Near Miss (NM)", "sum", "Leading Indicators", "int"),
    ("ncr_issued", "HSE NCRs Issued", "sum", "Leading Indicators", "int"),
    ("ncr_closed", "HSE NCRs Closed", "sum", "Leading Indicators", "int"),
    ("ncr_closure_rate", "NCRs Closure Rate", "last", "Leading Indicators", "percent"),
    ("tbt_sessions", "Toolbox Talks Sessions(TBT)", "sum", "Leading Indicators", "int"),
    ("hse_campaigns", "HSE Campaigns", "sum", "Leading Indicators", "int"),
    ("drills", "Drills", "sum", "Leading Indicators", "int"),
    ("mass_tbts", "Mass TBTs", "sum", "Leading Indicators", "int"),
    ("stop_work_notices", "Stop Work Notice Issued", "sum", "Leading Indicators", "int"),
    ("hse_rewards", "HSE Rewards (# of People)", "sum", "Leading Indicators", "int"),
    ("subcon_safety_meetings", "No. SubCon Safety meeting with CM", "sum", "Leading Indicators", "int"),
    ("violations_issued", "Voilations Issued", "sum", "Leading Indicators", "int"),
    ("health_surveillance_count", "Health Survalience Conducted (# of People)", "sum", "Leading Indicators", "int"),
    ("internal_audits", "Internal HSE Audits", "sum", "Inspections and Audits", "int"),
    ("external_audits", "External Audits", "sum", "Inspections and Audits", "int"),
    ("safety_mgmt_walkthroughs", "Safety Management Walkthroughs", "sum", "Inspections and Audits", "int"),
    ("site_walkthroughs", "No. of Site Walkthrough (CM/Site Eng/HSE Team)", "sum", "Inspections and Audits", "int"),
    ("safety_committee_meetings", "Project Safety Committee Meeting", "sum", "Inspections and Audits", "int"),
    ("safety_committee_closure_pct", "% Safety Committee Meeting & action closure", "avg", "Inspections and Audits", "percent"),
    ("sor_issued", "Safety Observations Issued(SOR)", "sum", "Inspections and Audits", "int"),
    ("sor_closed", "Safety Observations Clossed(SOR)", "sum", "Inspections and Audits", "int"),
    ("sor_closeout_ratio", "SOR Closeout Ratio", "last", "Inspections and Audits", "percent"),
    ("hse_inductions_people", "HSE Inductions(NO of People)", "sum", "OHS Trainings", "int"),
    ("hse_induction_compliance_pct", "HSE Inductions Compliance Percentage", "avg", "OHS Trainings", "percent"),
    ("training_sessions", "Internal / Third Party HSE Trainings (Sessions)", "sum", "OHS Trainings", "int"),
    ("training_manhours", "Training Manhours", "sum", "OHS Trainings", "float"),
    ("ptw_issued", "Permit to Work Issued (PTW)", "sum", "Permits", "int"),
    ("ptw_compliance_pct", "PTW Compliance Percentage", "avg", "Permits", "percent"),
]

DB_COLUMNS = [col for col, *_ in COLUMNS]

# A workbook missing any of these can't be treated as this template at all.
# Everything else in COLUMNS is optional per-import, so a slightly different
# revision of the sheet (a column added or dropped) still imports rather
# than being rejected outright.
REQUIRED_HEADERS = [
    "Total BEC Manpower",
    "Total Sub. Con Manpower",
    "Working Hours / Day",
    "Total Man Hours",
]

_HEADER_LOOKUP = {header.strip().lower(): col for col, header, *_ in COLUMNS}
_REQUIRED_COLS = {_HEADER_LOOKUP[h.strip().lower()] for h in REQUIRED_HEADERS}

_META_LABELS = {
    "project name:": "project_name",
    "responsible:": "responsible",
    "position:": "position",
    "email:": "email",
}


def _normalize(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().lower()


def _coerce_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def _coerce_number(value: Any, kind: str) -> float | int:
    if isinstance(value, bool):
        raise ValueError("boolean value")
    if isinstance(value, (int, float)):
        num = float(value)
    elif isinstance(value, str):
        text = value.strip()
        if not text:
            raise ValueError("empty value")
        is_percent_text = text.endswith("%")
        cleaned = text.rstrip("%").replace(",", "").strip()
        num = float(cleaned)
        if is_percent_text:
            num = num / 100
    else:
        raise ValueError(f"unsupported type {type(value).__name__}")
    return int(round(num)) if kind == "int" else num


def _extract_meta(worksheet: Any) -> dict[str, str]:
    meta: dict[str, str] = {}
    try:
        max_col = min(worksheet.max_column or 1, 40)
        for row_idx in range(1, min(worksheet.max_row or 1, 4) + 1):
            for col in range(1, max_col + 1):
                label = _normalize(worksheet.cell(row=row_idx, column=col).value)
                key = _META_LABELS.get(label)
                if key and col + 1 <= max_col:
                    value = worksheet.cell(row=row_idx, column=col + 1).value
                    if value not in (None, ""):
                        meta[key] = str(value).strip()
    except Exception:  # best-effort only - never blocks an otherwise valid import
        pass
    return meta


def _find_header_row(worksheet: Any) -> tuple[int, dict[int, str]] | None:
    """Scan the first rows of a sheet for a header row containing 'Date' plus
    a plausible number of the workbook's known column headers. Returns
    (row_index, {column_index: db_column}) or None if this sheet doesn't
    look like the HSE statistics template."""
    max_row = min(worksheet.max_row or 1, 15)
    max_col = min(worksheet.max_column or 1, 60)
    for row_idx in range(1, max_row + 1):
        row_values = [worksheet.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]
        normalized = [_normalize(v) for v in row_values]
        if "date" not in normalized:
            continue
        date_col = normalized.index("date") + 1
        col_map: dict[int, str] = {}
        for col, norm in enumerate(normalized, start=1):
            if norm in _HEADER_LOOKUP:
                col_map[col] = _HEADER_LOOKUP[norm]
        if len(col_map) >= 4:
            col_map[date_col] = "record_date"
            return row_idx, col_map
    return None


def parse_workbook(data: bytes) -> tuple[dict[str, str], list[dict[str, Any]], list[str], str]:
    """Parse an uploaded .xlsx file into daily HSE statistics rows.

    Returns (meta, rows, rejected_reasons, sheet_name). Raises
    ValidationError for anything that means the file can't be imported at
    all (corrupt file, no matching sheet, missing required columns, no data
    rows). Per-row problems (a bad date, a non-numeric cell) are collected
    into rejected_reasons and the row is skipped rather than failing the
    whole import.
    """
    if not data:
        raise ValidationError("The uploaded file is empty.")
    if len(data) < 4 or data[:2] != b"PK":
        raise ValidationError("The file is not a valid .xlsx workbook (not a recognizable Excel file).")

    try:
        # Not read_only: this module does random cell access (cell-by-cell
        # header detection, then re-reading earlier rows for the meta block),
        # which forces a read-only worksheet to re-run its streaming XML
        # parser from the start on every backward seek - unusably slow on a
        # sheet with dozens of columns. Workbooks here are small enough
        # (a few hundred KB) that loading fully into memory is cheap.
        workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=False)
    except (InvalidFileException, zipfile.BadZipFile, KeyError, OSError) as error:
        raise ValidationError("The workbook could not be read - it may be corrupted or password-protected.") from error

    sheet_name = None
    header_row_idx = None
    col_map: dict[int, str] = {}
    for name in workbook.sheetnames:
        found = _find_header_row(workbook[name])
        if found:
            header_row_idx, col_map = found
            sheet_name = name
            break

    if not sheet_name:
        raise ValidationError(
            "Could not find a worksheet with the expected HSE statistics header row "
            "(a 'Date' column alongside the tracked metrics, e.g. 'Total BEC Manpower')."
        )

    mapped_cols = set(col_map.values())
    missing = _REQUIRED_COLS - mapped_cols
    if missing:
        missing_headers = [header for col, header, *_ in COLUMNS if col in missing]
        raise ValidationError("Missing required column(s): " + ", ".join(missing_headers))

    worksheet = workbook[sheet_name]
    meta = _extract_meta(worksheet)

    rows: list[dict[str, Any]] = []
    rejected: list[str] = []
    max_row = worksheet.max_row or header_row_idx

    for row_idx in range(header_row_idx + 1, max_row + 1):
        raw = {db_col: worksheet.cell(row=row_idx, column=col).value for col, db_col in col_map.items()}
        record_date = raw.get("record_date")
        if record_date in (None, ""):
            continue

        parsed_date = _coerce_date(record_date)
        if parsed_date is None:
            rejected.append(f"Row {row_idx}: invalid date ({record_date!r})")
            continue

        # "Total BEC Manpower" is the workbook's own signal for whether a day
        # was actually filled in: several other columns are formulas that
        # resolve to 0 even on an untouched future/template day, so checking
        # "every cell is blank" doesn't reliably catch them - manpower is
        # always the first field entered by hand and stays genuinely blank
        # until someone does. REQUIRED_HEADERS guarantees this column exists.
        if raw.get("bec_manpower") in (None, ""):
            continue

        record: dict[str, Any] = {"record_date": parsed_date}
        bad_cell = None
        for col, header, _agg, _section, kind in COLUMNS:
            value = raw.get(col)
            if value in (None, ""):
                record[col] = None
                continue
            try:
                record[col] = _coerce_number(value, kind)
            except (TypeError, ValueError):
                bad_cell = f"Row {row_idx}: invalid value for '{header}' ({value!r})"
                break
        if bad_cell:
            rejected.append(bad_cell)
            continue

        rows.append(record)

    if not rows and not rejected:
        raise ValidationError("No data rows were found below the header row.")

    return meta, rows, rejected, sheet_name
