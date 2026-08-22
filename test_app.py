import csv
import io
import os
import tempfile
import unittest
import zipfile
from datetime import datetime, timedelta
from unittest.mock import patch
from zoneinfo import ZoneInfo

import openpyxl


TEST_FILES = tempfile.TemporaryDirectory()
os.environ["OHS_DB_PATH"] = os.path.join(TEST_FILES.name, "test.db")
os.environ["ADMIN_PASSWORD"] = "test-admin-password"
os.environ["SECRET_KEY"] = "test-secret-key-that-is-only-used-by-the-automated-suite"
os.environ["EXPORT_TOKEN"] = "test-export-token"

from app import CHECKLIST_ITEMS, app, database, sql  # noqa: E402


class ChecklistApplicationTests(unittest.TestCase):
    def setUp(self):
        app.config.update(TESTING=True, SESSION_COOKIE_SECURE=False)
        self.client = app.test_client()
        with database() as connection:
            connection.execute("DELETE FROM inspections")
            connection.execute("DELETE FROM near_miss_reports")
            connection.execute("DELETE FROM violation_notices")
            connection.execute("DELETE FROM ptw_logs")
            connection.execute("DELETE FROM hse_daily_stats")
            connection.execute("DELETE FROM hse_import_history")

    def payload(self):
        return {
            "projectName": "1 Hotel Diriyah",
            "workLocation": "Zone 3",
            "contractor": "Test Contractor",
            "inspectedBy": "QA Inspector",
            "inspectionDate": "2026-08-15",
            "inspectionTime": "14:30",
            "shift": "Day",
            "remarks": "Automated verification record.",
            "signoffName": "QA Inspector",
            "signed": True,
            "responses": {item["id"]: "Y" for item in CHECKLIST_ITEMS},
            "responseNotes": {},
        }

    def login(self):
        return self.client.post("/admin", data={"password": "test-admin-password"})

    def near_miss_payload(self):
        return {
            "departmentProject": "Zone 3",
            "incidentDate": "2026-08-16",
            "incidentTime": "10:00",
            "location": "Podium Level 2",
            "reportedBy": "Foreman A",
            "whatHappened": "Ladder slipped on a wet floor.",
            "nearMissTypes": ["Unsafe Condition"],
            "reportedBySignoff": "Foreman A",
        }

    def violation_payload(self):
        return {
            "projectName": "1 Hotel Diriyah",
            "violationDate": "2026-08-16",
            "employeeName": "John Doe",
            "companyContractor": "BEC Arabia Contracting",
            "violationLocation": "Zone 2",
            "violationType": "No PPE",
            "violationDescription": "Worker observed without a hard hat in an active work zone.",
            "actions": ["First Warning"],
            "issuedByName": "Site HSE Officer",
        }

    def ptw_payload(self):
        # A day out (rather than a fixed date) so this stays "open" under the
        # auto-close-on-expiry feature no matter when the suite actually runs.
        tomorrow = (datetime.now(ZoneInfo("Asia/Riyadh")) + timedelta(days=1)).strftime("%Y-%m-%d")
        return {
            "ptwNumber": "BAJV-829",
            "ptwType": "Hot work",
            "issuer": "Faisal",
            "receiver": "Sayed",
            "company": "BAJV",
            "location": "Basement",
            "workDescription": "Drilling, grinding & pipe installation",
            "startDate": tomorrow,
            "startTime": "08:00",
            "endDate": tomorrow,
            "endTime": "17:00",
        }

    def test_homepage_links_to_all_systems(self):
        response = self.client.get("/")
        self.assertEqual(response.status_code, 200)
        for path in ("/inspection", "/near-miss", "/violation", "/ptw", "/admin"):
            self.assertIn(path.encode(), response.data)

    def test_inspection_form_moved_from_root(self):
        response = self.client.get("/inspection")
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"OHS Team Inspection Checklist", response.data)

    def test_checklist_has_all_source_requirements(self):
        response = self.client.get("/api/checklist")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json["total"], 102)
        self.assertEqual(len(response.json["sections"]), 14)

    def test_submit_review_and_export_record(self):
        response = self.client.post("/api/inspections", json=self.payload())
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.json["score"], 100.0)
        self.assertEqual(response.json["reportNo"], "OHS-001")
        record_id = response.json["id"]

        self.assertEqual(self.login().status_code, 302)
        dashboard = self.client.get("/admin")
        self.assertIn(b"OHS-001", dashboard.data)
        detail = self.client.get(f"/admin/records/{record_id}")
        self.assertEqual(detail.status_code, 200)
        self.assertIn(b"GENERAL BEST PRACTICE", detail.data)

        summary = self.client.get("/admin/export?kind=summary")
        self.assertEqual(summary.status_code, 200)
        self.assertIn("attachment", summary.headers["Content-Disposition"])
        self.assertIn("OHS-001", summary.get_data(as_text=True))

        detailed = self.client.get("/admin/export?kind=detailed")
        rows = list(csv.reader(io.StringIO(detailed.get_data(as_text=True).lstrip("\ufeff"))))
        self.assertEqual(len(rows), 103)
        self.assertEqual(rows[-1][6], "General Best Practice")

    def test_non_compliance_requires_observation(self):
        payload = self.payload()
        first = CHECKLIST_ITEMS[0]
        payload["responses"][first["id"]] = "N"
        response = self.client.post("/api/inspections", json=payload)
        self.assertEqual(response.status_code, 400)
        self.assertIn("observation", response.json["error"].lower())

    def test_admin_requires_correct_password(self):
        response = self.client.get("/admin/export")
        self.assertEqual(response.status_code, 302)
        response = self.client.post("/admin", data={"password": "wrong"})
        self.assertIn(b"Incorrect admin password", response.data)

    def test_trends_tab_renders_with_and_without_data(self):
        self.login()
        empty = self.client.get("/admin?view=trends")
        self.assertEqual(empty.status_code, 200)
        self.assertIn(b"<svg", empty.data)

        self.client.post("/api/inspections", json=self.payload())
        with_data = self.client.get("/admin?view=trends")
        self.assertEqual(with_data.status_code, 200)
        self.assertIn(b"<svg", with_data.data)

    def test_trends_legend_uses_css_classes_not_inline_style(self):
        # Inline style="" attributes are silently dropped by the strict
        # style-src 'self' CSP — the legend color swatches must use CSS
        # classes instead, or they render invisible with no error shown.
        self.login()
        response = self.client.get("/admin?view=trends")
        self.assertNotIn(b'<i style=', response.data)

    def test_near_miss_form_pages_load(self):
        self.assertEqual(self.client.get("/near-miss").status_code, 200)
        self.assertEqual(self.client.get("/violation").status_code, 200)
        self.assertEqual(self.client.get("/ptw").status_code, 200)

    def test_near_miss_requires_at_least_one_type(self):
        payload = self.near_miss_payload()
        payload["nearMissTypes"] = []
        response = self.client.post("/api/near-miss", json=payload)
        self.assertEqual(response.status_code, 400)

    def test_submit_review_and_export_near_miss(self):
        response = self.client.post("/api/near-miss", json=self.near_miss_payload())
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.json["reportNo"], "NEAR-MISS-001")
        record_id = response.json["id"]

        self.login()
        dashboard = self.client.get("/admin?view=near-miss")
        self.assertIn(b"NEAR-MISS-001", dashboard.data)
        detail = self.client.get(f"/admin/near-miss/{record_id}")
        self.assertEqual(detail.status_code, 200)
        self.assertIn(b"NEAR MISS REPORTING FORM", detail.data)
        self.assertIn(b"Unsafe Condition", detail.data)

        export = self.client.get("/admin/export/near-miss")
        self.assertEqual(export.status_code, 200)
        self.assertIn("NEAR-MISS-001", export.get_data(as_text=True))

    def test_violation_requires_employee_name(self):
        payload = self.violation_payload()
        payload["employeeName"] = ""
        response = self.client.post("/api/violations", json=payload)
        self.assertEqual(response.status_code, 400)

    def test_submit_review_and_export_violation(self):
        response = self.client.post("/api/violations", json=self.violation_payload())
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.json["violationNo"], "VIOLATION-001")
        record_id = response.json["id"]

        self.login()
        dashboard = self.client.get("/admin?view=violations")
        self.assertIn(b"VIOLATION-001", dashboard.data)
        detail = self.client.get(f"/admin/violations/{record_id}")
        self.assertEqual(detail.status_code, 200)
        self.assertIn(b"VIOLATION NOTICE", detail.data.upper())
        self.assertIn(b"First Warning", detail.data)

        export = self.client.get("/admin/export/violations")
        self.assertEqual(export.status_code, 200)
        self.assertIn("VIOLATION-001", export.get_data(as_text=True))

    def test_ptw_requires_valid_type(self):
        payload = self.ptw_payload()
        payload["ptwType"] = "Not a real type"
        response = self.client.post("/api/ptw", json=payload)
        self.assertEqual(response.status_code, 400)

    def test_ptw_requires_ptw_number(self):
        payload = self.ptw_payload()
        payload["ptwNumber"] = ""
        response = self.client.post("/api/ptw", json=payload)
        self.assertEqual(response.status_code, 400)

    def test_ptw_form_suggests_next_permit_number(self):
        blank = self.client.get("/ptw")
        self.assertIn(b'value="BAJV-1"', blank.data)

        payload = self.ptw_payload()
        payload["ptwNumber"] = "BAJV-840"
        self.client.post("/api/ptw", json=payload)

        after = self.client.get("/ptw")
        self.assertIn(b'value="BAJV-841"', after.data)

    def test_submit_review_and_export_ptw(self):
        response = self.client.post("/api/ptw", json=self.ptw_payload())
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.json["ptwNumber"], "BAJV-829")
        record_id = response.json["id"]

        self.login()
        dashboard = self.client.get("/admin?view=ptw")
        self.assertIn(b"BAJV-829", dashboard.data)
        detail = self.client.get(f"/admin/ptw/{record_id}")
        self.assertEqual(detail.status_code, 200)
        self.assertIn(b"BAJV-829", detail.data)
        # New entries default to open until a coordinator closes them out.
        self.assertIn(b"open", detail.data.lower())

        export = self.client.get("/admin/export/ptw")
        self.assertEqual(export.status_code, 200)
        self.assertIn("BAJV-829", export.get_data(as_text=True))

    def test_ptw_xlsx_export_matches_original_columns_and_highlights_open(self):
        open_id = self.client.post("/api/ptw", json=self.ptw_payload()).json["id"]
        closed_payload = self.ptw_payload()
        closed_payload["ptwNumber"] = "BAJV-830"
        closed_id = self.client.post("/api/ptw", json=closed_payload).json["id"]

        self.login()
        self.client.post(f"/admin/ptw/{closed_id}", data={
            "ptwNumber": "BAJV-830", "issuer": "Faisal", "receiver": "Sayed", "ptwType": "Hot work",
            "workDescription": "Drilling", "areaHsePersonnel": "", "location": "Basement", "shift": "",
            "startDate": "2026-08-18", "startTime": "08:00", "endDate": "2026-08-18", "endTime": "17:00",
            "company": "BAJV", "status": "closed", "workersCount": "", "reviewedBy": "",
        })

        response = self.client.get("/admin/export/ptw.xlsx")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.content_type,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        workbook = openpyxl.load_workbook(io.BytesIO(response.data))
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]
        self.assertEqual(header, [
            "S.N", "PTW Number", "PTW Issuer", "PTW Receiver", "Type of PTW", "Work Description",
            "Area HSE Personnel", "Location", "Shift", "PTW Start Date & Time", "PTW End Date & Time",
            "Company Name", "Status", "No. of Workers", "Reviewed By",
        ])

        rows = {row[1].value: row for row in sheet.iter_rows(min_row=2)}
        self.assertIn("BAJV-829", rows)
        self.assertIn("BAJV-830", rows)
        # Open permit row is highlighted; the closed one is not.
        self.assertEqual(rows["BAJV-829"][1].fill.fgColor.rgb, "00FFFF00")
        self.assertNotEqual(rows["BAJV-830"][1].fill.fgColor.rgb, "00FFFF00")

    def test_ptw_auto_closes_once_end_time_has_passed(self):
        yesterday = (datetime.now(ZoneInfo("Asia/Riyadh")) - timedelta(days=1)).strftime("%Y-%m-%d")
        payload = self.ptw_payload()
        payload["ptwNumber"] = "BAJV-900"
        payload["startDate"] = yesterday
        payload["endDate"] = yesterday
        record_id = self.client.post("/api/ptw", json=payload).json["id"]

        with database() as connection:
            row = connection.execute(sql("SELECT status FROM ptw_logs WHERE id = ?"), [record_id]).fetchone()
            self.assertEqual(dict(row)["status"], "open")

        self.login()
        # Merely loading the list (or overview, export, etc.) is what triggers
        # the self-correction — there's no separate scheduled job to wait for.
        self.client.get("/admin?view=ptw")

        with database() as connection:
            row = connection.execute(sql("SELECT status FROM ptw_logs WHERE id = ?"), [record_id]).fetchone()
            self.assertEqual(dict(row)["status"], "closed")

    def test_ptw_list_orders_by_permit_number_not_submission_order(self):
        # Submitted out of numeric order (834 first, then 832, then 833) to
        # simulate someone backfilling or correcting an entry after the fact.
        for number in ("BAJV-834", "BAJV-832", "BAJV-833"):
            payload = self.ptw_payload()
            payload["ptwNumber"] = number
            response = self.client.post("/api/ptw", json=payload)
            self.assertEqual(response.status_code, 201)

        self.login()
        dashboard = self.client.get("/admin?view=ptw")
        body = dashboard.data.decode()
        # Highest number first, regardless of submission order.
        self.assertLess(body.index("BAJV-834"), body.index("BAJV-833"))
        self.assertLess(body.index("BAJV-833"), body.index("BAJV-832"))

    def test_ptw_overview_breaks_down_open_permits_by_area_and_type(self):
        # Two open in Basement (one Hot work, one Lifting), one open in Zone A
        # (Hot work), and one closed in Basement that should be excluded entirely.
        basement_hot = self.ptw_payload()
        self.client.post("/api/ptw", json=basement_hot)

        basement_lift = self.ptw_payload()
        basement_lift.update({"ptwNumber": "BAJV-830", "ptwType": "Lifting"})
        self.client.post("/api/ptw", json=basement_lift)

        zone_a_hot = self.ptw_payload()
        zone_a_hot.update({"ptwNumber": "BAJV-831", "location": "Zone A"})
        self.client.post("/api/ptw", json=zone_a_hot)

        closed_payload = self.ptw_payload()
        closed_payload["ptwNumber"] = "BAJV-832"
        closed_id = self.client.post("/api/ptw", json=closed_payload).json["id"]
        self.login()
        self.client.post(f"/admin/ptw/{closed_id}", data={
            "ptwNumber": "BAJV-832", "issuer": "Faisal", "receiver": "Sayed", "ptwType": "Hot work",
            "workDescription": "Drilling", "areaHsePersonnel": "", "location": "Basement", "shift": "",
            "startDate": "2026-08-18", "startTime": "08:00", "endDate": "2026-08-18", "endTime": "17:00",
            "company": "BAJV", "status": "closed", "workersCount": "", "reviewedBy": "",
        })

        overview = self.client.get("/admin?view=ptw")
        self.assertEqual(overview.status_code, 200)
        body = overview.data.decode()
        # 3 open total (the closed one excluded); Basement busiest with 2; Hot work count is 2.
        self.assertIn("Basement", body)
        self.assertIn("Hot work (1)", body)
        self.assertIn("Lifting (1)", body)

    def test_ptw_edit_updates_status_and_persists(self):
        record_id = self.client.post("/api/ptw", json=self.ptw_payload()).json["id"]
        self.login()
        payload = self.ptw_payload()
        payload["status"] = "closed"
        payload["reviewedBy"] = "Faisal"
        response = self.client.post(f"/admin/ptw/{record_id}", data={
            "ptwNumber": payload["ptwNumber"], "issuer": payload["issuer"], "receiver": payload["receiver"],
            "ptwType": payload["ptwType"], "workDescription": payload["workDescription"], "areaHsePersonnel": "",
            "location": payload["location"], "shift": "", "startDate": payload["startDate"],
            "startTime": payload["startTime"], "endDate": payload["endDate"], "endTime": payload["endTime"],
            "company": payload["company"], "status": "closed", "workersCount": "", "reviewedBy": "Faisal",
        })
        self.assertEqual(response.status_code, 302)
        detail = self.client.get(f"/admin/ptw/{record_id}")
        self.assertIn(b'value="Faisal"', detail.data)
        export = self.client.get("/admin/export/ptw")
        self.assertIn("closed", export.get_data(as_text=True))

    def test_ptw_edit_rejects_invalid_update_and_keeps_entered_values(self):
        record_id = self.client.post("/api/ptw", json=self.ptw_payload()).json["id"]
        self.login()
        response = self.client.post(f"/admin/ptw/{record_id}", data={
            "ptwNumber": "", "issuer": "Faisal", "receiver": "Sayed", "ptwType": "Hot work",
            "workDescription": "Drilling", "areaHsePersonnel": "", "location": "Basement", "shift": "",
            "startDate": "2026-08-18", "startTime": "08:00", "endDate": "2026-08-18", "endTime": "17:00",
            "company": "BAJV", "status": "open", "workersCount": "", "reviewedBy": "",
        })
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"PTW number is required", response.data)

    def test_delete_ptw(self):
        record_id = self.client.post("/api/ptw", json=self.ptw_payload()).json["id"]
        self.login()
        response = self.client.post(f"/admin/ptw/{record_id}/delete")
        self.assertEqual(response.status_code, 302)
        detail = self.client.get(f"/admin/ptw/{record_id}")
        self.assertEqual(detail.status_code, 404)

    def test_sequential_report_numbers(self):
        first = self.client.post("/api/inspections", json=self.payload())
        ignored_payload = self.payload()
        ignored_payload["reportNo"] = "CUSTOM-IGNORED"
        second = self.client.post("/api/inspections", json=ignored_payload)
        self.assertEqual(first.json["reportNo"], "OHS-001")
        self.assertEqual(second.json["reportNo"], "OHS-002")

    def test_photo_upload_rejects_bad_token(self):
        response = self.client.post("/api/uploads/bad token!", data={}, content_type="multipart/form-data")
        self.assertEqual(response.status_code, 400)

    def test_photo_upload_without_storage_configured(self):
        response = self.client.post(
            "/api/uploads/abcdef1234567890",
            data={"photo": (io.BytesIO(b"not-a-real-image"), "photo.jpg")},
            content_type="multipart/form-data",
        )
        self.assertEqual(response.status_code, 503)

    def test_b2_client_uses_path_style_and_disables_chunked_checksums(self):
        with patch.dict(os.environ, {
            "B2_ENDPOINT": "s3.us-west-004.backblazeb2.com",
            "B2_KEY_ID": "test-key-id",
            "B2_APPLICATION_KEY": "test-app-key",
        }):
            from app import b2_client
            client = b2_client()
        config = client.meta.config
        self.assertEqual(config.s3["addressing_style"], "path")
        self.assertEqual(config.request_checksum_calculation, "when_required")
        self.assertEqual(config.response_checksum_validation, "when_required")
        self.assertEqual(client.meta.region_name, "us-west-004")

    def test_over_max_content_length_returns_json_not_html_on_api_routes(self):
        payload = self.near_miss_payload()
        payload["whatHappened"] = "x" * (13 * 1024 * 1024)
        response = self.client.post("/api/near-miss", json=payload)
        self.assertEqual(response.status_code, 413)
        self.assertEqual(response.content_type, "application/json")
        self.assertIn("error", response.json)

    def test_delete_requires_admin(self):
        record_id = self.client.post("/api/inspections", json=self.payload()).json["id"]
        response = self.client.post(f"/admin/records/{record_id}/delete")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/admin", response.headers["Location"])
        self.login()
        detail = self.client.get(f"/admin/records/{record_id}")
        self.assertEqual(detail.status_code, 200)

    def test_delete_record(self):
        record_id = self.client.post("/api/inspections", json=self.payload()).json["id"]
        self.login()
        response = self.client.post(f"/admin/records/{record_id}/delete")
        self.assertEqual(response.status_code, 302)
        detail = self.client.get(f"/admin/records/{record_id}")
        self.assertEqual(detail.status_code, 404)

    def test_delete_near_miss(self):
        record_id = self.client.post("/api/near-miss", json=self.near_miss_payload()).json["id"]
        self.login()
        response = self.client.post(f"/admin/near-miss/{record_id}/delete")
        self.assertEqual(response.status_code, 302)
        detail = self.client.get(f"/admin/near-miss/{record_id}")
        self.assertEqual(detail.status_code, 404)

    def test_delete_violation(self):
        record_id = self.client.post("/api/violations", json=self.violation_payload()).json["id"]
        self.login()
        response = self.client.post(f"/admin/violations/{record_id}/delete")
        self.assertEqual(response.status_code, 302)
        detail = self.client.get(f"/admin/violations/{record_id}")
        self.assertEqual(detail.status_code, 404)

    def test_near_miss_and_violation_detail_pages_render(self):
        near_miss_id = self.client.post("/api/near-miss", json=self.near_miss_payload()).json["id"]
        violation_id = self.client.post("/api/violations", json=self.violation_payload()).json["id"]
        self.login()
        near_miss_detail = self.client.get(f"/admin/near-miss/{near_miss_id}")
        self.assertEqual(near_miss_detail.status_code, 200)
        violation_detail = self.client.get(f"/admin/violations/{violation_id}")
        self.assertEqual(violation_detail.status_code, 200)

    def test_detail_pages_survive_malformed_legacy_json(self):
        near_miss_id = self.client.post("/api/near-miss", json=self.near_miss_payload()).json["id"]
        violation_id = self.client.post("/api/violations", json=self.violation_payload()).json["id"]
        with database() as connection:
            connection.execute(
                sql("UPDATE near_miss_reports SET near_miss_types = ?, photos = ? WHERE id = ?"),
                ["not-json", "not-json", near_miss_id],
            )
            connection.execute(
                sql("UPDATE violation_notices SET actions = ?, photos = ? WHERE id = ?"),
                ["not-json", "not-json", violation_id],
            )
        self.login()
        near_miss_detail = self.client.get(f"/admin/near-miss/{near_miss_id}")
        self.assertEqual(near_miss_detail.status_code, 200)
        violation_detail = self.client.get(f"/admin/violations/{violation_id}")
        self.assertEqual(violation_detail.status_code, 200)

    def test_backup_rejects_missing_or_wrong_token(self):
        self.assertEqual(self.client.get("/admin/backup").status_code, 401)
        self.assertEqual(self.client.get("/admin/backup?token=wrong").status_code, 401)

    def test_backup_returns_zip_with_all_record_types(self):
        self.client.post("/api/inspections", json=self.payload())
        self.client.post("/api/near-miss", json=self.near_miss_payload())
        self.client.post("/api/violations", json=self.violation_payload())
        self.client.post("/api/ptw", json=self.ptw_payload())

        response = self.client.get("/admin/backup?token=test-export-token")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, "application/zip")

        archive = zipfile.ZipFile(io.BytesIO(response.data))
        names = archive.namelist()
        self.assertEqual(len(names), 5)
        self.assertTrue(any(name.startswith("inspections-summary-") for name in names))
        self.assertTrue(any(name.startswith("inspections-detailed-") for name in names))
        self.assertTrue(any(name.startswith("near-miss-") for name in names))
        self.assertTrue(any(name.startswith("violations-") for name in names))
        self.assertTrue(any(name.startswith("ptw-log-") for name in names))


def build_hse_workbook(rows: list[tuple]) -> bytes:
    """rows: list of (date_str, bec, subcon, hours, man_hours) tuples."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Statistics"
    sheet.append([
        "Date", "Total BEC Manpower", "Total Sub. Con Manpower", "Working Hours / Day",
        "Total Man Hours", "Toolbox Talks Sessions(TBT)", "Number of Fatalities",
    ])
    for date_str, bec, subcon, hours, man_hours in rows:
        sheet.append([date_str, bec, subcon, hours, man_hours, 1, 0])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class HseStatisticsTests(unittest.TestCase):
    def setUp(self):
        from app import app as flask_app
        flask_app.config.update(TESTING=True, SESSION_COOKIE_SECURE=False)
        self.client = flask_app.test_client()
        with database() as connection:
            connection.execute("DELETE FROM hse_daily_stats")
            connection.execute("DELETE FROM hse_import_history")

    def login(self):
        return self.client.post("/admin", data={"password": "test-admin-password"})

    def upload(self, data: bytes, filename: str = "stats.xlsx"):
        return self.client.post(
            "/admin/hse/upload",
            data={"workbook": (io.BytesIO(data), filename)},
            content_type="multipart/form-data",
        )

    def test_upload_requires_admin_session(self):
        response = self.upload(build_hse_workbook([("2026-05-01", 10, 5, 10, 150)]))
        self.assertEqual(response.status_code, 302)

    def test_dashboard_and_status_api_require_admin_session(self):
        self.assertEqual(self.client.get("/api/hse/dashboard").status_code, 302)
        self.assertEqual(self.client.get("/api/hse/status").status_code, 302)

    def test_valid_workbook_imports_and_appears_on_dashboard(self):
        self.login()
        data = build_hse_workbook([
            ("2026-05-01", 10, 5, 10, 150),
            ("2026-05-02", 12, 6, 10, 180),
        ])
        response = self.upload(data)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json["status"], "success")
        self.assertEqual(response.json["recordsInserted"], 2)
        self.assertEqual(response.json["recordsUpdated"], 0)
        self.assertEqual(response.json["recordsRejected"], 0)

        dashboard = self.client.get("/api/hse/dashboard")
        self.assertEqual(dashboard.status_code, 200)
        records = dashboard.json["records"]
        self.assertEqual(len(records), 2)
        self.assertEqual(records[0]["date"], "2026-05-01")
        self.assertEqual(records[0]["Total BEC Manpower"], 10)
        self.assertEqual(dashboard.json["totalDays"], 2)

    def test_reupload_updates_in_place_without_duplicating(self):
        self.login()
        first = build_hse_workbook([("2026-05-01", 10, 5, 10, 150)])
        self.upload(first)

        second = build_hse_workbook([("2026-05-01", 20, 8, 10, 280)])  # same date, new numbers
        response = self.upload(second)
        self.assertEqual(response.json["recordsInserted"], 0)
        self.assertEqual(response.json["recordsUpdated"], 1)

        dashboard = self.client.get("/api/hse/dashboard")
        records = dashboard.json["records"]
        self.assertEqual(len(records), 1)  # no duplicate row for the same date
        self.assertEqual(records[0]["Total BEC Manpower"], 20)  # latest values won

    def test_rejects_non_xlsx_filename(self):
        self.login()
        response = self.upload(b"whatever", filename="stats.csv")
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json["status"], "failed")

    def test_rejects_corrupted_file_without_touching_existing_data(self):
        self.login()
        good = build_hse_workbook([("2026-05-01", 10, 5, 10, 150)])
        self.upload(good)

        response = self.upload(b"not a real xlsx file at all", filename="corrupt.xlsx")
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json["status"], "failed")

        # the previously imported row must still be intact
        dashboard = self.client.get("/api/hse/dashboard")
        self.assertEqual(len(dashboard.json["records"]), 1)

    def test_rejects_workbook_missing_required_columns(self):
        # Enough recognized headers to be detected as *a* statistics sheet
        # (so this exercises the "missing required column" path, not the
        # separate "no matching sheet at all" path), but none of them are
        # in REQUIRED_HEADERS.
        self.login()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Date", "Near Miss (NM)", "Toolbox Talks Sessions(TBT)", "HSE NCRs Issued", "Drills"])
        sheet.append(["2026-05-01", 1, 2, 0, 0])
        buffer = io.BytesIO()
        workbook.save(buffer)

        response = self.upload(buffer.getvalue())
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json["status"], "failed")
        self.assertIn("Missing required column", response.json["error"])

    def test_rejects_workbook_with_no_recognizable_header_row(self):
        self.login()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Date", "Some Unrelated Column"])
        sheet.append(["2026-05-01", 5])
        buffer = io.BytesIO()
        workbook.save(buffer)

        response = self.upload(buffer.getvalue())
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json["status"], "failed")
        self.assertIn("Could not find a worksheet", response.json["error"])

    def test_import_history_is_recorded_for_success_and_failure(self):
        self.login()
        self.upload(build_hse_workbook([("2026-05-01", 10, 5, 10, 150)]), filename="good.xlsx")
        self.upload(b"garbage", filename="bad.xlsx")

        with database() as connection:
            cursor = connection.cursor()
            cursor.execute("SELECT filename, status FROM hse_import_history ORDER BY seq")
            history = [dict(row) for row in cursor.fetchall()]
        self.assertEqual(len(history), 2)
        self.assertEqual(history[0]["filename"], "good.xlsx")
        self.assertEqual(history[0]["status"], "success")
        self.assertEqual(history[1]["filename"], "bad.xlsx")
        self.assertEqual(history[1]["status"], "failed")

    def test_admin_page_shows_hse_tabs(self):
        self.login()
        response = self.client.get("/admin?view=hse-stats")
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"HSE Statistics", response.data)

        response = self.client.get("/admin?view=hse-upload")
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Data Management", response.data)


if __name__ == "__main__":
    unittest.main()
